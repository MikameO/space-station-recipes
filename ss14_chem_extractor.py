"""
SS14 Chemistry Database Extractor
Fetches all reagent/reaction YAML prototypes from SS14 + RMC14 repos,
parses them, builds craft dependency trees, and generates an Excel file.
"""

import json
import os
import re
import time
import urllib.request
import urllib.error
from pathlib import Path
from collections import defaultdict

import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    FORK_REGISTRY,
    VANILLA_RAW, RMC14_RAW,
    VANILLA_REAGENT_FILES, VANILLA_REACTION_FILES, VANILLA_LOCALE_FILES,
    RMC14_REAGENT_FILES, RMC14_REACTION_FILES, RMC14_LOCALE_FILES,
    VANILLA_SEED_FILES, RMC14_SEED_FILES, VANILLA_BOTANY_LOCALE_FILES,
    OTHER_REAGENT_SOURCES, DANGEROUS_INTERACTIONS,
    RMC14_BLOCKED_REACTIONS, RMC14_MODIFIED_REACTIONS,
    BASE_DISPENSER_CHEMICALS, CATEGORY_SHEET_MAP,
    REAGENT_COLUMNS, REACTION_COLUMNS,
    ANTAG_DATA, ANTAG_STRATEGIES, DELIVERY_MECHANISMS, SYNDICATE_ITEMS,
)

# ─────────────────────────────────────────────
# Fork source detection (generalized from hardcoded _RMC14 check)
# ─────────────────────────────────────────────

def detect_fork_source(source_file: str) -> str:
    """Determine which fork a file belongs to based on its path.
    Returns fork ID from FORK_REGISTRY, or 'vanilla' if no match."""
    for fork_id, config in FORK_REGISTRY.items():
        if fork_id == "vanilla":
            continue
        custom_dir = config.get("custom_dir")
        if custom_dir and custom_dir in source_file:
            return fork_id
    return "vanilla"

SCRIPT_DIR = Path(__file__).parent
CACHE_DIR = SCRIPT_DIR / "cache"
OUTPUT_FILE = SCRIPT_DIR / "SS14_Chemistry_Database.xlsx"
JSON_FILE = SCRIPT_DIR / "data.json"


# ─────────────────────────────────────────────
# Auto-diff: compare fork's vanilla-path reactions against originals
# ─────────────────────────────────────────────

def compare_reaction(vanilla_rxn: dict, fork_rxn: dict) -> list[str]:
    """Compare two parsed reaction dicts. Returns list of human-readable diff strings."""
    diffs = []

    # Compare reactants
    v_reacts = vanilla_rxn.get("reactants", {})
    f_reacts = fork_rxn.get("reactants", {})
    v_rids = set(v_reacts.keys()) if isinstance(v_reacts, dict) else set()
    f_rids = set(f_reacts.keys()) if isinstance(f_reacts, dict) else set()

    for rid in f_rids - v_rids:
        amt = f_reacts[rid].get("amount", 1) if isinstance(f_reacts[rid], dict) else f_reacts[rid]
        diffs.append(f"Added {rid}({amt}) reactant")
    for rid in v_rids - f_rids:
        diffs.append(f"Removed {rid} reactant")
    for rid in v_rids & f_rids:
        v_amt = v_reacts[rid].get("amount", 1) if isinstance(v_reacts[rid], dict) else v_reacts[rid]
        f_amt = f_reacts[rid].get("amount", 1) if isinstance(f_reacts[rid], dict) else f_reacts[rid]
        if v_amt != f_amt:
            diffs.append(f"{rid} amount {v_amt}->{f_amt}")
        v_cat = v_reacts[rid].get("catalyst", False) if isinstance(v_reacts[rid], dict) else False
        f_cat = f_reacts[rid].get("catalyst", False) if isinstance(f_reacts[rid], dict) else False
        if v_cat != f_cat:
            diffs.append(f"{rid} catalyst {'added' if f_cat else 'removed'}")

    # Compare products
    v_prods = vanilla_rxn.get("products", {})
    f_prods = fork_rxn.get("products", {})
    for pid in set(f_prods) - set(v_prods):
        diffs.append(f"Added product {pid}")
    for pid in set(v_prods) - set(f_prods):
        diffs.append(f"Removed product {pid}")
    for pid in set(v_prods) & set(f_prods):
        if v_prods[pid] != f_prods[pid]:
            diffs.append(f"{pid} yield {v_prods[pid]}->{f_prods[pid]}")

    # Compare temperature
    for temp_key, label in [("minTemp", "minTemp"), ("maxTemp", "maxTemp")]:
        v_t = vanilla_rxn.get(temp_key)
        f_t = fork_rxn.get(temp_key)
        if v_t != f_t:
            if v_t is None and f_t is not None:
                diffs.append(f"Added {label}: {f_t}")
            elif v_t is not None and f_t is None:
                diffs.append(f"Removed {label}")
            else:
                diffs.append(f"{label} {v_t}->{f_t}")

    # Compare mixer requirements
    v_mix = vanilla_rxn.get("requiredMixerCategories", [])
    f_mix = fork_rxn.get("requiredMixerCategories", [])
    if v_mix != f_mix:
        if not v_mix and f_mix:
            diffs.append(f"Added mixer: {', '.join(f_mix)}")
        elif v_mix and not f_mix:
            diffs.append("Removed mixer requirement")
        else:
            diffs.append(f"Mixer changed: {v_mix}->{f_mix}")

    # Compare priority
    v_pri = vanilla_rxn.get("priority")
    f_pri = fork_rxn.get("priority")
    if v_pri != f_pri and (v_pri is not None or f_pri is not None):
        diffs.append(f"Priority {v_pri}->{f_pri}")

    return diffs


def diff_fork_reactions(
    vanilla_rxns: dict, fork_vanilla_rxns: dict,
    manual_blocked: set, manual_modified: dict,
    blocked_categories: set | None = None,
    all_reagents: dict | None = None,
) -> tuple[set, dict]:
    """Compare fork's vanilla-path reactions against originals.
    Returns (blocked_ids, modified_dict). Manual overrides take precedence.
    blocked_categories: if set, block all vanilla reactions whose products
    belong to these categories (e.g. RMC14 replaces all Medicine)."""
    blocked = set(manual_blocked)
    modified = dict(manual_modified)

    for rid, v_rxn in vanilla_rxns.items():
        if rid in manual_blocked or rid in manual_modified:
            continue  # Manual override takes precedence

        # Category-level blocking (e.g. RMC14 replaces entire Medicine category)
        if blocked_categories and all_reagents:
            products = v_rxn.get("products", {})
            if any(
                all_reagents.get(pid, {}).get("category", "") in blocked_categories
                for pid in products
            ):
                blocked.add(rid)
                continue

        if rid not in fork_vanilla_rxns:
            # Reaction removed/commented out in fork
            blocked.add(rid)
        else:
            f_rxn = fork_vanilla_rxns[rid]
            diffs = compare_reaction(v_rxn, f_rxn)
            if diffs:
                modified[rid] = "; ".join(diffs)

    return blocked, modified


# ─────────────────────────────────────────────
# Phase 1: Data Fetching
# ─────────────────────────────────────────────

def fetch_file(url: str, cache_path: Path) -> str:
    """Download a file from URL with local caching."""
    if cache_path.exists():
        return cache_path.read_text(encoding="utf-8")

    cache_path.parent.mkdir(parents=True, exist_ok=True)
    print(f"  Fetching: {url.split('master/')[-1]}")
    for attempt in range(3):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "SS14-Chem-Extractor/1.0"})
            with urllib.request.urlopen(req, timeout=30) as resp:
                content = resp.read().decode("utf-8")
            cache_path.write_text(content, encoding="utf-8")
            time.sleep(0.3)
            return content
        except urllib.error.HTTPError as e:
            if e.code in (404, 403):
                print(f"  WARNING: HTTP {e.code} for {url}")
                return ""  # permanent error, no retry
            if attempt < 2:
                print(f"  HTTP {e.code}, retry {attempt + 1}/3...")
                time.sleep(2 * (attempt + 1))
                continue
            print(f"  FAILED: HTTP {e.code} after 3 attempts for {url}")
            return ""
        except Exception as e:
            if attempt < 2:
                print(f"  Retry {attempt + 1}/3: {e}")
                time.sleep(2 * (attempt + 1))
            else:
                print(f"  FAILED after 3 attempts: {e} for {url}")
                return ""
    return ""


def fetch_all_files(file_list: list[str], url_template: str, cache_subdir: str) -> dict[str, str]:
    """Fetch all files from a manifest list. Returns {path: content}."""
    results = {}
    for path in file_list:
        url = url_template.format(path=path)
        cache_path = CACHE_DIR / cache_subdir / path.replace("/", os.sep)
        content = fetch_file(url, cache_path)
        if content:
            results[path] = content
    return results


def fetch_and_extract_sprites():
    """Fetch RSI sprite PNGs from SS14 repo and save as individual files.

    RSI format: each state is a separate PNG file inside the .rsi directory.
    We fetch {state}.png directly and optionally scale 2x for crisp display.
    """
    from config import SPRITE_MANIFEST

    sprites_dir = SCRIPT_DIR / "sprites"
    sprites_dir.mkdir(exist_ok=True)

    try:
        from PIL import Image
    except ImportError:
        print("  WARNING: Pillow not installed — skipping sprite extraction")
        print("  Install with: pip install Pillow")
        return

    base_url = VANILLA_RAW.replace("{path}", "Resources/Textures/{rsi}/{state}.png")
    fetched = 0

    for entry in SPRITE_MANIFEST:
        sprite_id = entry["id"]
        rsi_path = entry["rsi"]
        state = entry["state"]
        out_path = sprites_dir / f"{sprite_id}.png"

        if out_path.exists():
            fetched += 1
            continue

        url = base_url.format(rsi=rsi_path, state=state)
        print(f"  Fetching sprite: {sprite_id} ({rsi_path}/{state}.png)")

        try:
            req = urllib.request.Request(url, headers={"User-Agent": "SS14-Chem-Extractor/1.0"})
            with urllib.request.urlopen(req, timeout=15) as resp:
                raw = resp.read()

            # Open, scale 2x for crisp pixel-art display, save
            import io
            img = Image.open(io.BytesIO(raw))
            # Only take a 32x32 crop from top-left for multi-direction states
            w, h = img.size
            if w > 32 or h > 32:
                img = img.crop((0, 0, min(w, 32), min(h, 32)))
            # Scale 2x with nearest-neighbor (preserves pixel art)
            img = img.resize((64, 64), Image.NEAREST)
            img.save(out_path, "PNG")
            fetched += 1
            time.sleep(0.3)
        except Exception as e:
            print(f"  WARNING: Failed to fetch sprite {sprite_id}: {e}")

    print(f"  Sprites: {fetched}/{len(SPRITE_MANIFEST)} extracted to {sprites_dir}")


# ─────────────────────────────────────────────
# Phase 2: YAML Parsing
# ─────────────────────────────────────────────

class SS14Loader(yaml.SafeLoader):
    """Custom YAML loader that handles SS14's !type: tags without mutating global SafeLoader."""
    pass


def _type_constructor(loader_inst, suffix, node):
    """Handle !type:SomeType tags by converting to dict with _type key."""
    if isinstance(node, yaml.MappingNode):
        data = loader_inst.construct_mapping(node, deep=True)
        data["_type"] = suffix
        return data
    elif isinstance(node, yaml.ScalarNode):
        val = loader_inst.construct_scalar(node)
        return {"_type": suffix, "_value": val}
    elif isinstance(node, yaml.SequenceNode):
        val = loader_inst.construct_sequence(node, deep=True)
        return {"_type": suffix, "_items": val}
    return {"_type": suffix}


SS14Loader.add_multi_constructor("!type:", _type_constructor)


def make_yaml_loader():
    """Return the SS14-aware YAML loader class."""
    return SS14Loader


def parse_yaml_content(content: str, source_name: str, loader) -> list[dict]:
    """Parse YAML content and return list of prototype entries."""
    try:
        docs = list(yaml.load_all(content, Loader=loader))
    except yaml.YAMLError as e:
        print(f"  YAML error in {source_name}: {e}")
        return []

    entries = []
    for doc in docs:
        if isinstance(doc, list):
            for item in doc:
                if isinstance(item, dict):
                    item["_source_file"] = source_name
                    entries.append(item)
        elif isinstance(doc, dict):
            doc["_source_file"] = source_name
            entries.append(doc)
    return entries


def parse_all_prototypes(files: dict[str, str], loader) -> tuple[dict, dict]:
    """Parse all YAML files into reagent and reaction dictionaries."""
    reagents = {}
    reactions = {}

    for path, content in files.items():
        entries = parse_yaml_content(content, path, loader)
        for entry in entries:
            proto_type = entry.get("type", "")
            proto_id = entry.get("id", "")
            if not proto_id:
                continue
            if proto_type == "reagent":
                reagents[proto_id] = entry
            elif proto_type == "reaction":
                reactions[proto_id] = entry
    return reagents, reactions


# ─────────────────────────────────────────────
# Phase 3: Localization
# ─────────────────────────────────────────────

def parse_ftl_content(content: str) -> dict[str, str]:
    """Parse Fluent (.ftl) file into key-value dict."""
    locale = {}
    current_key = None
    current_val_lines = []

    for line in content.split("\n"):
        stripped = line.strip()
        if not stripped or stripped.startswith("#"):
            if current_key and current_val_lines:
                locale[current_key] = " ".join(current_val_lines).strip()
                current_key = None
                current_val_lines = []
            continue

        # Attribute pattern: .attr = value
        if current_key and stripped.startswith("."):
            # Save previous key
            if current_val_lines:
                locale[current_key] = " ".join(current_val_lines).strip()

            attr_match = re.match(r"\.(\w[\w-]*)\s*=\s*(.*)", stripped)
            if attr_match:
                attr_name = attr_match.group(1)
                attr_val = attr_match.group(2).strip()
                # Store as parent-key.attr-name
                parent = current_key.rsplit(".", 1)[0] if "." in current_key else current_key
                locale[f"{parent}.{attr_name}"] = attr_val
                current_key = f"{parent}.{attr_name}"
                current_val_lines = [attr_val] if attr_val else []
            continue

        # Main key = value pattern
        match = re.match(r"([\w][\w-]*(?:\.[\w-]+)*)\s*=\s*(.*)", line)
        if match:
            if current_key and current_val_lines:
                locale[current_key] = " ".join(current_val_lines).strip()
            current_key = match.group(1)
            val = match.group(2).strip()
            current_val_lines = [val] if val else []
        elif current_key and (line.startswith("    ") or line.startswith("\t")):
            # Continuation line
            current_val_lines.append(stripped)

    if current_key and current_val_lines:
        locale[current_key] = " ".join(current_val_lines).strip()

    return locale


def load_all_localization(files: dict[str, str]) -> dict[str, str]:
    """Merge all FTL files into a single locale dictionary."""
    locale = {}
    for path, content in files.items():
        parsed = parse_ftl_content(content)
        locale.update(parsed)
    # Strip Fluent placeables: {$var}, {-term}, {FUNC()} from values
    for key in locale:
        locale[key] = re.sub(r'\{[^}]+\}', '', locale[key]).strip()
    return locale


def pascal_to_words(name: str) -> str:
    """Convert PascalCase to 'Pascal Case'."""
    words = re.sub(r"(?<=[a-z])(?=[A-Z])|(?<=[A-Z])(?=[A-Z][a-z])", " ", name)
    return words


def resolve_name(reagent: dict, locale: dict) -> str:
    """Get human-readable name for a reagent."""
    name_key = reagent.get("name", "")
    if name_key and name_key in locale:
        return locale[name_key]
    # Try by ID
    rid = reagent.get("id", "")
    for pattern in [f"reagent-name-{rid.lower()}", f"reagent-name-{re.sub(r'(?<!^)(?=[A-Z])', '-', rid).lower()}"]:
        if pattern in locale:
            return locale[pattern]
    return pascal_to_words(rid) if rid else name_key


def resolve_desc(reagent: dict, locale: dict) -> str:
    """Get description for a reagent."""
    desc_key = reagent.get("desc", "")
    if desc_key and desc_key in locale:
        return locale[desc_key]
    rid = reagent.get("id", "")
    for pattern in [f"reagent-desc-{rid.lower()}", f"reagent-desc-{re.sub(r'(?<!^)(?=[A-Z])', '-', rid).lower()}"]:
        if pattern in locale:
            return locale[pattern]
    return ""


# ─────────────────────────────────────────────
# Phase 4: Parent Resolution & Inheritance
# ─────────────────────────────────────────────

def deep_merge(base: dict, override: dict) -> dict:
    """Deep merge two dicts. Override wins on conflicts."""
    result = base.copy()
    for key, val in override.items():
        if key in result and isinstance(result[key], dict) and isinstance(val, dict):
            result[key] = deep_merge(result[key], val)
        else:
            result[key] = val
    return result


def resolve_parents(reagents: dict) -> dict:
    """Resolve parent inheritance for all reagents."""
    resolved = {}
    resolving = set()  # cycle detection

    def resolve_one(rid: str) -> dict:
        if rid in resolved:
            return resolved[rid]
        if rid not in reagents:
            return {}
        if rid in resolving:
            return reagents[rid]  # break cycle

        resolving.add(rid)
        entry = reagents[rid]
        parent = entry.get("parent")

        if not parent:
            resolved[rid] = entry
            resolving.discard(rid)
            return entry

        parents = parent if isinstance(parent, list) else [parent]
        merged = {}
        for p in parents:
            parent_data = resolve_one(p)
            merged = deep_merge(merged, parent_data)

        # These fields must NOT be inherited from parents
        for no_inherit in ("abstract", "parent", "type", "id"):
            merged.pop(no_inherit, None)

        # Child overrides parent, but keep _source_file from child
        source = entry.get("_source_file", "")
        result = deep_merge(merged, entry)
        result["_source_file"] = source
        resolved[rid] = result
        resolving.discard(rid)
        return result

    for rid in reagents:
        resolve_one(rid)

    return resolved


# ─────────────────────────────────────────────
# Phase 5: Dependency Trees
# ─────────────────────────────────────────────

def build_reaction_lookup(reactions: dict) -> dict[str, list[dict]]:
    """Build product -> [reaction] lookup."""
    lookup = defaultdict(list)
    for rid, rxn in reactions.items():
        products = rxn.get("products", {})
        for product_id in products:
            lookup[product_id].append(rxn)
    return dict(lookup)


def identify_base_chemicals(reagents: dict, reaction_lookup: dict) -> set[str]:
    """Identify base chemicals (dispenser + never-produced)."""
    base = set(BASE_DISPENSER_CHEMICALS)
    for rid in reagents:
        if rid not in reaction_lookup:
            base.add(rid)
    return base


def get_craft_chain_compact(reagent_id: str, reaction_lookup: dict, base_set: set,
                            visited: set | None = None) -> str:
    """Build compact one-line craft chain string."""
    if visited is None:
        visited = set()
    if reagent_id in visited:
        return f"{reagent_id}[LOOP]"
    if reagent_id in base_set or reagent_id not in reaction_lookup:
        return f"{reagent_id}[B]"

    visited = visited | {reagent_id}
    rxn = reaction_lookup[reagent_id][0]  # take first reaction path
    reactants = rxn.get("reactants", {})

    parts = []
    for reactant_id, info in reactants.items():
        amount = info.get("amount", 1) if isinstance(info, dict) else 1
        catalyst = info.get("catalyst", False) if isinstance(info, dict) else False
        sub = get_craft_chain_compact(reactant_id, reaction_lookup, base_set, visited)
        prefix = f"{amount}x" if amount != 1 else ""
        cat_mark = "[cat]" if catalyst else ""
        parts.append(f"{prefix}{sub}{cat_mark}")

    return f"{reagent_id}({' + '.join(parts)})"


def get_craft_chain_expanded(reagent_id: str, reaction_lookup: dict, base_set: set,
                              indent: int = 0, visited: set | None = None) -> str:
    """Build expanded multi-line craft chain."""
    if visited is None:
        visited = set()
    prefix = "  " * indent

    if reagent_id in visited:
        return f"{prefix}{reagent_id} [CIRCULAR]\n"
    if reagent_id in base_set or reagent_id not in reaction_lookup:
        return f"{prefix}{reagent_id} [BASE]\n"

    visited = visited | {reagent_id}
    rxn = reaction_lookup[reagent_id][0]
    reactants = rxn.get("reactants", {})
    products = rxn.get("products", {})

    prod_amount = products.get(reagent_id, "?")
    temp_str = ""
    if rxn.get("minTemp"):
        temp_str += f" (min {rxn['minTemp']}K)"
    if rxn.get("maxTemp"):
        temp_str += f" (max {rxn['maxTemp']}K)"
    mixer = rxn.get("requiredMixerCategories", [])
    mixer_str = f" [{', '.join(mixer)}]" if mixer else ""

    lines = [f"{prefix}{reagent_id} = {prod_amount}u{temp_str}{mixer_str}\n"]
    for reactant_id, info in reactants.items():
        amount = info.get("amount", 1) if isinstance(info, dict) else 1
        catalyst = info.get("catalyst", False) if isinstance(info, dict) else False
        cat_mark = " (CATALYST)" if catalyst else ""
        lines.append(f"{prefix}  + {amount}x {reactant_id}{cat_mark}\n")
        if reactant_id not in base_set and reactant_id in reaction_lookup and reactant_id not in visited:
            lines.append(get_craft_chain_expanded(reactant_id, reaction_lookup, base_set, indent + 2, visited))

    return "".join(lines)


# ─────────────────────────────────────────────
# Phase 6: Effect Summarization
# ─────────────────────────────────────────────

def summarize_single_effect(effect: dict) -> str:
    """Summarize a single metabolism effect entry."""
    if not isinstance(effect, dict):
        return str(effect)

    etype = effect.get("_type", effect.get("type", "Unknown"))
    parts = []

    if "HealthChange" in etype:
        damage = effect.get("damage", {})
        if isinstance(damage, dict):
            groups = damage.get("groups", {}) if isinstance(damage.get("groups"), dict) else {}
            types_ = damage.get("types", {}) if isinstance(damage.get("types"), dict) else {}
            # Some HealthChange effects have damage as flat dict {DamageType: amount}
            flat = {k: v for k, v in damage.items() if k not in ("groups", "types", "_type", "_value") and isinstance(v, (int, float))}
            all_dmg = {**groups, **types_, **flat}
            even = "Even" in etype
            for name, val in all_dmg.items():
                direction = "Heals" if val < 0 else "Deals"
                even_mark = " (even)" if even else ""
                parts.append(f"{direction} {name} {abs(val)}{even_mark}")
        if not parts:
            parts.append("EvenHealthChange" if "Even" in etype else "HealthChange")

    elif "SatiateThirst" in etype:
        factor = effect.get("factor", 1)
        parts.append(f"Thirst x{factor}")

    elif "SatiateHunger" in etype:
        factor = effect.get("factor", 1)
        parts.append(f"Hunger x{factor}")

    elif "MovementSpeedModifier" in etype:
        walk = effect.get("walkSpeedModifier", 1)
        sprint = effect.get("sprintSpeedModifier", 1)
        parts.append(f"Speed walk={walk} sprint={sprint}")

    elif "ModifyStatusEffect" in etype or "GenericStatusEffect" in etype:
        key = effect.get("key", effect.get("statusEffectId", etype))
        parts.append(f"Status: {key}")

    elif "Emote" in etype:
        emote = effect.get("emote", "?")
        parts.append(f"Emote: {emote}")

    elif "Jitter" in etype:
        parts.append("Jitter")

    elif "Drunk" in etype:
        parts.append("Drunk")

    elif "PopupMessage" in etype:
        parts.append("Message popup")

    elif "AdjustReagent" in etype:
        reagent = effect.get("reagent", "?")
        amount = effect.get("amount", 0)
        parts.append(f"{'Adds' if amount > 0 else 'Removes'} {abs(amount)}u {reagent}")

    elif "AdjustTemperature" in etype:
        amount = effect.get("amount", 0)
        parts.append(f"Temp {'+'if amount > 0 else ''}{amount}")

    elif "Vomit" in etype or "ChemVomit" in etype:
        parts.append("Vomit")

    elif "CureDisease" in etype:
        parts.append("Cures disease")

    elif "Oxygenate" in etype:
        parts.append("Oxygenates blood")

    elif "ModifyBloodLevel" in etype:
        parts.append(f"Blood level {effect.get('amount', '?')}")

    elif "ModifyBleedAmount" in etype:
        parts.append(f"Bleed {effect.get('amount', '?')}")

    elif "ResetNarcolepsy" in etype:
        parts.append("Resets narcolepsy")

    elif "Electrocute" in etype:
        parts.append("Electrocution")

    elif "FlammableReaction" in etype or "Flammable" in etype:
        parts.append("Flammable")

    elif "Paralyze" in etype:
        parts.append("Paralyze")

    elif "CreateGas" in etype:
        parts.append(f"Creates gas: {effect.get('gas', '?')}")

    elif "Explosion" in etype or "ExplosionReactionEffect" in etype:
        parts.append("Explosion")

    elif "CreateEntityReactionEffect" in etype or "SpawnEntity" in etype:
        parts.append(f"Spawns: {effect.get('entity', '?')}")

    elif "PlantMetabolism" in etype or "PlantAdjust" in etype:
        parts.append(f"Plant effect")

    else:
        # Fallback: just name the type
        short = etype.replace("ReagentEffect", "").replace("Reaction", "")
        parts.append(f"Effect: {short}")

    # Add conditions if present
    conditions = effect.get("conditions", [])
    if conditions:
        cond_parts = []
        for cond in conditions:
            if isinstance(cond, dict):
                ct = cond.get("_type", cond.get("type", ""))
                if "ReagentThreshold" in ct:
                    min_v = cond.get("min", "")
                    max_v = cond.get("max", "")
                    if min_v:
                        cond_parts.append(f">{min_v}u")
                    if max_v:
                        cond_parts.append(f"<{max_v}u")
                elif "Temperature" in ct:
                    cond_parts.append(f"temp cond")
                else:
                    cond_parts.append(ct.split(":")[-1] if ":" in ct else ct)
        if cond_parts:
            parts.append(f"(if {', '.join(cond_parts)})")

    prob = effect.get("probability", None)
    if prob is not None and prob != 1:
        parts.append(f"@{int(prob * 100)}%")

    return " ".join(parts) if parts else etype


def extract_effects_list(value) -> list[dict]:
    """Extract flat list of effect dicts from various metabolism structures."""
    results = []
    if isinstance(value, list):
        for item in value:
            results.extend(extract_effects_list(item))
    elif isinstance(value, dict):
        # If it has an 'effects' key, those are the real effects
        inner = value.get("effects")
        if inner is not None:
            results.extend(extract_effects_list(inner))
        elif "_type" in value:
            # This dict IS an effect
            results.append(value)
        else:
            # Might be a plain dict of effects by name
            for k, v in value.items():
                if k.startswith("_"):
                    continue
                results.extend(extract_effects_list(v))
    return results


def summarize_metabolisms(metabolisms: dict) -> tuple[str, str]:
    """Summarize all metabolisms into (effects_text, metabolism_paths)."""
    if not metabolisms or not isinstance(metabolisms, dict):
        return "", ""

    all_effects = []
    paths = []

    for path_name, path_value in metabolisms.items():
        if path_name.startswith("_"):
            continue
        paths.append(path_name)
        effects = extract_effects_list(path_value)
        for eff in effects:
            summary = summarize_single_effect(eff)
            if summary:
                all_effects.append(f"[{path_name}] {summary}")

    return "; ".join(all_effects), ", ".join(paths)


def extract_effect_tags(metabolisms: dict) -> list[str]:
    """Extract searchable effect type tags from metabolisms for filtering."""
    if not metabolisms or not isinstance(metabolisms, dict):
        return []
    tags = set()
    for path, value in metabolisms.items():
        if path.startswith("_"):
            continue
        for eff in extract_effects_list(value):
            etype = eff.get("_type", "")
            damage = eff.get("damage", {})
            if "HealthChange" in etype and isinstance(damage, dict):
                groups = damage.get("groups", {}) if isinstance(damage.get("groups"), dict) else {}
                types_ = damage.get("types", {}) if isinstance(damage.get("types"), dict) else {}
                flat = {k: v for k, v in damage.items()
                        if k not in ("groups", "types", "_type", "_value") and isinstance(v, (int, float))}
                for name, val in {**groups, **types_, **flat}.items():
                    tag = name.lower()
                    if val < 0:
                        tags.add(f"heals:{tag}")
                    elif val > 0:
                        tags.add(f"deals:{tag}")
            elif "SatiateThirst" in etype:
                tags.add("thirst")
            elif "SatiateHunger" in etype:
                tags.add("hunger")
            elif "Drunk" in etype:
                tags.add("drunk")
            elif "Jitter" in etype:
                tags.add("jitter")
            elif "Vomit" in etype or "ChemVomit" in etype:
                tags.add("vomit")
            elif "Emote" in etype:
                tags.add("emote")
            elif "Oxygenate" in etype:
                tags.add("heals:airloss")
            elif "ModifyBloodLevel" in etype:
                tags.add("blood")
            elif "ModifyBleedAmount" in etype:
                tags.add("bleed")
            elif "CureDisease" in etype or "CureZombieInfection" in etype:
                tags.add("cure")
            elif "Explosion" in etype:
                tags.add("explosion")
            elif "Flammable" in etype:
                tags.add("flammable")
            elif "Paralyze" in etype:
                tags.add("paralyze")
            elif "MovementSpeedModifier" in etype:
                tags.add("speed")
            elif "AdjustTemperature" in etype:
                tags.add("temperature")
    return sorted(tags)


def extract_overdose_threshold(metabolisms: dict) -> float | None:
    """Extract overdose threshold from metabolism conditions (vanilla SS14 pattern)."""
    if not metabolisms or not isinstance(metabolisms, dict):
        return None
    thresholds = []
    for path, value in metabolisms.items():
        if path.startswith("_"):
            continue
        for eff in extract_effects_list(value):
            conditions = eff.get("conditions", [])
            if not isinstance(conditions, list):
                continue
            for cond in conditions:
                if not isinstance(cond, dict):
                    continue
                ct = cond.get("_type", "")
                if "ReagentThreshold" in ct or "ReagentCondition" in ct:
                    min_val = cond.get("min")
                    if min_val and isinstance(min_val, (int, float)) and min_val > 0:
                        # Check if this condition gates a NEGATIVE effect (damage, jitter, vomit)
                        etype = eff.get("_type", "")
                        damage = eff.get("damage", {})
                        is_negative = (
                            "Jitter" in etype or "Vomit" in etype or "Drunk" in etype
                            or "Paralyze" in etype
                            or (isinstance(damage, dict) and any(
                                v > 0 for k, v in {
                                    **damage.get("groups", {}), **damage.get("types", {}),
                                    **{k2: v2 for k2, v2 in damage.items()
                                       if isinstance(v2, (int, float))}
                                }.items()
                            ))
                        )
                        if is_negative:
                            thresholds.append(min_val)
    return min(thresholds) if thresholds else None


def summarize_reaction_effects(reaction: dict) -> str:
    """Summarize reaction-specific effects (not metabolisms)."""
    effects = reaction.get("effects", [])
    if not effects:
        return ""
    parts = []
    for eff in effects:
        if isinstance(eff, dict):
            parts.append(summarize_single_effect(eff))
    return "; ".join(parts)


# ─────────────────────────────────────────────
# Phase 7: Categorization
# ─────────────────────────────────────────────

def categorize_reagent(reagent: dict) -> str:
    """Determine which sheet category a reagent belongs to."""
    source = reagent.get("_source_file", "")
    group = reagent.get("group", "")

    # Fork-specific reagents get "{ForkName} SubCategory" sheets
    fork_id = detect_fork_source(source)
    if fork_id != "vanilla":
        fork_name = FORK_REGISTRY[fork_id]["name"]
        src_lower = source.lower()
        if "medicine" in src_lower or group == "Medicine":
            return f"{fork_name} Medicine"
        elif "toxin" in src_lower or group == "Toxins":
            return f"{fork_name} Toxins"
        elif "element" in src_lower or group == "Elements":
            return f"{fork_name} Elements"
        elif "drink" in src_lower:
            return f"{fork_name} Drinks"
        elif "narcotic" in src_lower or group == "Narcotics":
            return f"{fork_name} Narcotics"
        elif "pyrotechnic" in src_lower or group == "Pyrotechnic":
            return f"{fork_name} Pyrotechnic"
        else:
            return f"{fork_name} Other"

    # Vanilla categorization by source file path
    if "Drink/alcohol" in source:
        return "Drinks (Alcoholic)"
    if "Drink/" in source:
        return "Drinks (Non-Alc)"
    if "Food/" in source:
        return "Food & Condiments"
    if "Materials/" in source:
        return "Materials"

    # By group field
    group_map = {
        "Elements": "Elements",
        "Medicine": "Medicine",
        "Chemicals": "Chemicals",
        "Toxins": "Toxins",
        "Narcotics": "Narcotics",
        "Cleaning": "Cleaning",
        "Pyrotechnic": "Pyrotechnic",
        "Gases": "Gases",
        "Botany": "Botany",
        "Biological": "Biological",
        "Fun": "Fun",
        "Alcohol": "Drinks (Alcoholic)",
        "Drinks": "Drinks (Non-Alc)",
    }
    if group in group_map:
        return group_map[group]

    # By source filename
    fname = source.split("/")[-1].replace(".yml", "")
    return CATEGORY_SHEET_MAP.get(fname.capitalize(), "Other")


def categorize_reaction(reaction: dict) -> str:
    """Determine sheet for a reaction."""
    source = reaction.get("_source_file", "")
    fork_id = detect_fork_source(source)
    if fork_id != "vanilla":
        return f"{FORK_REGISTRY[fork_id]['name']} Recipes"

    fname = source.split("/")[-1].replace(".yml", "")
    mapping = {
        "drinks": "Drink Recipes",
        "medicine": "All Reactions",
        "chemicals": "All Reactions",
        "biological": "All Reactions",
        "botany": "All Reactions",
        "cleaning": "All Reactions",
        "food": "All Reactions",
        "fun": "All Reactions",
        "gas": "All Reactions",
        "pyrotechnic": "All Reactions",
        "single_reagent": "All Reactions",
        "soap": "All Reactions",
    }
    return mapping.get(fname, "All Reactions")


# ─────────────────────────────────────────────
# Phase 8: Excel Generation
# ─────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def apply_headers(ws, columns: list[str]):
    """Write and format header row."""
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def safe_hex_fill(color_str: str) -> PatternFill | None:
    """Create a PatternFill from a hex color string, or None if invalid."""
    if not color_str:
        return None
    color = color_str.strip().lstrip("#")
    # Handle RGBA (take first 6 chars for RGB)
    if len(color) == 8:
        color = color[:6]
    if len(color) != 6:
        return None
    try:
        int(color, 16)
        return PatternFill(start_color=color, end_color=color, fill_type="solid")
    except ValueError:
        return None


def write_reagent_row(ws, row: int, reagent: dict, locale: dict,
                      reaction_lookup: dict, base_set: set):
    """Write a single reagent row."""
    rid = reagent.get("id", "")
    name = resolve_name(reagent, locale)
    group = reagent.get("group", "")
    color = reagent.get("color", "")

    # Find recipe (reaction that produces this reagent)
    recipe_str = ""
    produces_str = ""
    temp_str = ""
    mixer_str = ""
    if rid in reaction_lookup:
        rxn = reaction_lookup[rid][0]
        reactants = rxn.get("reactants", {})
        parts = []
        for react_id, info in reactants.items():
            amount = info.get("amount", 1) if isinstance(info, dict) else 1
            catalyst = info.get("catalyst", False) if isinstance(info, dict) else False
            cat = " (cat)" if catalyst else ""
            parts.append(f"{amount}x {react_id}{cat}")
        recipe_str = " + ".join(parts)

        products = rxn.get("products", {})
        prod_parts = [f"{v}x {k}" for k, v in products.items()]
        produces_str = ", ".join(prod_parts)

        min_t = rxn.get("minTemp")
        max_t = rxn.get("maxTemp")
        temp_parts = []
        if min_t:
            temp_parts.append(f"min {min_t}K")
        if max_t:
            temp_parts.append(f"max {max_t}K")
        temp_str = ", ".join(temp_parts)

        mixers = rxn.get("requiredMixerCategories", [])
        mixer_str = ", ".join(mixers) if mixers else ""

    # Effects
    metabolisms = reagent.get("metabolisms", {})
    effects_str, metabolism_paths = summarize_metabolisms(metabolisms)

    # Overdose — RMC14 top-level field + vanilla threshold extraction
    overdose = ""
    if reagent.get("overdose"):
        overdose = str(reagent["overdose"])
    else:
        od = extract_overdose_threshold(reagent.get("metabolisms", {}))
        if od:
            overdose = str(od)
    if reagent.get("criticalOverdose"):
        overdose += f" / crit: {reagent['criticalOverdose']}"

    # Craft chain
    craft_chain = ""
    if rid in reaction_lookup:
        try:
            craft_chain = get_craft_chain_compact(rid, reaction_lookup, base_set)
        except RecursionError:
            craft_chain = "[TOO DEEP]"

    # Physical description
    phys_key = reagent.get("physicalDesc", "")
    phys_desc = locale.get(phys_key, phys_key) if phys_key else ""

    # Flavor
    flavor = reagent.get("flavor", "")
    if isinstance(flavor, dict):
        flavor = str(flavor)

    # Source
    source = reagent.get("_source_file", "")
    source_label = FORK_REGISTRY.get(detect_fork_source(source), {}).get("name", "Vanilla")
    source_short = f"{source_label}: {source.split('/')[-1]}"

    # Notes
    notes_parts = []
    if reagent.get("abstract"):
        notes_parts.append("ABSTRACT")
    if reagent.get("slipData"):
        notes_parts.append("Slippery")
    if reagent.get("fizziness"):
        notes_parts.append(f"Fizzy: {reagent['fizziness']}")
    if reagent.get("worksOnTheDead"):
        notes_parts.append("Works on dead")
    if reagent.get("plantMetabolism"):
        notes_parts.append("Plant metabolism")
    notes_str = "; ".join(notes_parts)

    # Write cells
    values = [
        name, rid, group, color, recipe_str, produces_str,
        temp_str, mixer_str, effects_str, metabolism_paths, overdose,
        craft_chain, phys_desc, flavor, source_short, notes_str,
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=str(val) if val else "")
        cell.alignment = CELL_ALIGN
        cell.border = THIN_BORDER

    # Color fill for color column
    if color:
        fill = safe_hex_fill(color)
        if fill:
            ws.cell(row=row, column=4).fill = fill
            # Use white text if dark background
            try:
                r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
                if (r * 299 + g * 587 + b * 114) / 1000 < 128:
                    ws.cell(row=row, column=4).font = Font(color="FFFFFF")
            except (ValueError, IndexError):
                pass


def write_reaction_row(ws, row: int, reaction: dict, reaction_lookup: dict, base_set: set):
    """Write a single reaction row."""
    rid = reaction.get("id", "")
    reactants = reaction.get("reactants", {})
    products = reaction.get("products", {})

    react_parts = []
    for react_id, info in reactants.items():
        amount = info.get("amount", 1) if isinstance(info, dict) else 1
        catalyst = info.get("catalyst", False) if isinstance(info, dict) else False
        cat = " (catalyst)" if catalyst else ""
        react_parts.append(f"{amount}x {react_id}{cat}")

    prod_parts = [f"{v}x {k}" for k, v in products.items()]

    min_temp = reaction.get("minTemp", "")
    max_temp = reaction.get("maxTemp", "")
    mixers = reaction.get("requiredMixerCategories", [])
    effects = summarize_reaction_effects(reaction)
    priority = reaction.get("priority", "")
    impact = reaction.get("impact", "")

    # Craft chain for primary product
    craft_chain = ""
    if products:
        primary = list(products.keys())[0]
        try:
            craft_chain = get_craft_chain_compact(primary, reaction_lookup, base_set)
        except RecursionError:
            craft_chain = "[TOO DEEP]"

    source = reaction.get("_source_file", "")
    source_label = FORK_REGISTRY.get(detect_fork_source(source), {}).get("name", "Vanilla")
    source_short = f"{source_label}: {source.split('/')[-1]}"

    values = [
        rid, " + ".join(react_parts), ", ".join(prod_parts),
        str(min_temp) if min_temp else "", str(max_temp) if max_temp else "",
        ", ".join(mixers), effects, str(priority) if priority else "",
        str(impact) if impact else "", craft_chain, source_short,
    ]
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.alignment = CELL_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws, max_width: int = 50):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)


def generate_excel(reagents: dict, reactions: dict, locale: dict,
                   reaction_lookup: dict, base_set: set):
    """Generate the complete Excel workbook."""
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Group reagents by category
    reagent_groups = defaultdict(list)
    for rid, reagent in sorted(reagents.items()):
        if reagent.get("abstract"):
            continue
        cat = categorize_reagent(reagent)
        reagent_groups[cat].append(reagent)

    # Group reactions by category
    reaction_groups = defaultdict(list)
    for rid, reaction in sorted(reactions.items()):
        cat = categorize_reaction(reaction)
        reaction_groups[cat].append(reaction)
        # Also put non-drink reactions into All Reactions
        if cat != "All Reactions" and cat != "RMC14 Recipes" and cat != "Drink Recipes":
            reaction_groups["All Reactions"].append(reaction)

    # === Overview Sheet ===
    ws = wb.create_sheet("Overview")
    ws.cell(row=1, column=1, value="SS14 Chemistry Database").font = Font(bold=True, size=16)
    ws.cell(row=2, column=1, value=f"Generated: {time.strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=4, column=1, value="Category").font = Font(bold=True)
    ws.cell(row=4, column=2, value="Count").font = Font(bold=True)
    row = 5
    for cat in sorted(reagent_groups.keys()):
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=len(reagent_groups[cat]))
        row += 1
    ws.cell(row=row + 1, column=1, value="Total Reagents").font = Font(bold=True)
    ws.cell(row=row + 1, column=2, value=sum(len(v) for v in reagent_groups.values())).font = Font(bold=True)
    ws.cell(row=row + 3, column=1, value="Reaction Category").font = Font(bold=True)
    ws.cell(row=row + 3, column=2, value="Count").font = Font(bold=True)
    r2 = row + 4
    for cat in sorted(reaction_groups.keys()):
        ws.cell(row=r2, column=1, value=cat)
        ws.cell(row=r2, column=2, value=len(reaction_groups[cat]))
        r2 += 1

    ws.cell(row=r2 + 2, column=1, value="Legend").font = Font(bold=True)
    ws.cell(row=r2 + 3, column=1, value="[B] = Base dispenser chemical")
    ws.cell(row=r2 + 4, column=1, value="[cat] = Catalyst (not consumed)")
    ws.cell(row=r2 + 5, column=1, value="[LOOP] = Circular dependency")
    ws.cell(row=r2 + 6, column=1, value="Temp in Kelvin (273K = 0°C, 310K = 37°C)")
    auto_width(ws)

    # === Reagent Sheets ===
    sheet_order = [
        "Elements", "Medicine", "Chemicals", "Toxins", "Narcotics",
        "Cleaning", "Pyrotechnic", "Gases", "Botany", "Biological", "Fun",
        "Drinks (Alcoholic)", "Drinks (Non-Alc)", "Food & Condiments", "Materials",
        "RMC14 Medicine", "RMC14 Elements", "RMC14 Toxins", "RMC14 Drinks",
        "RMC14 Narcotics", "RMC14 Pyrotechnic", "RMC14 Other",
    ]

    for sheet_name in sheet_order:
        reagents_list = reagent_groups.get(sheet_name, [])
        if not reagents_list:
            continue

        # Excel sheet name max 31 chars
        ws_name = sheet_name[:31]
        ws = wb.create_sheet(ws_name)
        apply_headers(ws, REAGENT_COLUMNS)

        for idx, reagent in enumerate(reagents_list, 2):
            write_reagent_row(ws, idx, reagent, locale, reaction_lookup, base_set)
        auto_width(ws)

    # Also handle any uncategorized
    for sheet_name, reagents_list in reagent_groups.items():
        if sheet_name not in sheet_order and reagents_list:
            ws_name = sheet_name[:31]
            if ws_name not in wb.sheetnames:
                ws = wb.create_sheet(ws_name)
                apply_headers(ws, REAGENT_COLUMNS)
                for idx, reagent in enumerate(reagents_list, 2):
                    write_reagent_row(ws, idx, reagent, locale, reaction_lookup, base_set)
                auto_width(ws)

    # === Reaction Sheets ===
    reaction_order = ["Drink Recipes", "All Reactions", "RMC14 Recipes"]
    for sheet_name in reaction_order:
        reactions_list = reaction_groups.get(sheet_name, [])
        if not reactions_list:
            continue

        ws = wb.create_sheet(sheet_name[:31])
        apply_headers(ws, REACTION_COLUMNS)

        for idx, rxn in enumerate(reactions_list, 2):
            write_reaction_row(ws, idx, rxn, reaction_lookup, base_set)
        auto_width(ws)

    # === Craft Trees Sheet ===
    ws = wb.create_sheet("Craft Trees")
    ws.cell(row=1, column=1, value="Reagent").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Full Craft Tree").font = Font(bold=True)
    ws.freeze_panes = "A2"
    row = 2
    for rid in sorted(reaction_lookup.keys()):
        if rid in reagents and not reagents[rid].get("abstract"):
            try:
                tree = get_craft_chain_expanded(rid, reaction_lookup, base_set)
            except RecursionError:
                tree = f"{rid} [TOO DEEP]"
            ws.cell(row=row, column=1, value=rid)
            ws.cell(row=row, column=2, value=tree)
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
    auto_width(ws, max_width=100)

    return wb


# ─────────────────────────────────────────────
# Phase 8b: Plant/Seed Source Extraction
# ─────────────────────────────────────────────

def parse_seed_sources(seed_files: dict[str, str], loader,
                       locale: dict | None = None) -> dict[str, list[str]]:
    """Parse seed YAML files to build reagent -> [plant names] mapping."""
    reagent_plants = defaultdict(list)

    for path, content in seed_files.items():
        entries = parse_yaml_content(content, path, loader)
        for entry in entries:
            if entry.get("type") != "seed":
                continue
            raw_name = entry.get("name", entry.get("id", "unknown"))
            # Resolve locale name
            if locale and raw_name in locale:
                seed_name = locale[raw_name]
            else:
                # Clean up seed name: seeds-ambrosiavulgaris-name -> Ambrosia Vulgaris
                clean = raw_name.replace("seeds-", "").replace("-name", "").replace("-", " ")
                clean = re.sub(r'([a-z])([A-Z])', r'\1 \2', clean)
                # Fix known compound words that are all lowercase
                fixes = {
                    "ambrosiavulgaris": "Ambrosia Vulgaris", "ambrosiadeus": "Ambrosia Deus",
                    "flyamanita": "Fly Amanita", "deathnettle": "Death Nettle",
                    "galaxythistle": "Galaxy Thistle", "spacemans trumpet": "Spaceman's Trumpet",
                    "onionred": "Red Onion", "bluetomato": "Blue Tomato",
                    "bloodtomato": "Blood Tomato", "holymelon": "Holy Melon",
                    "goldenapple": "Golden Apple", "bluepumpkin": "Blue Pumpkin",
                    "rainbow cannabis": "Rainbow Cannabis",
                    "walkingmushroom": "Walking Mushroom",
                }
                seed_name = fixes.get(clean.lower().strip(), clean.title())
            chemicals = entry.get("chemicals", {})
            if not isinstance(chemicals, dict):
                continue
            for reagent_id in chemicals:
                if reagent_id in ("Nutriment", "Vitamin", "Fiber"):
                    continue
                plant_label = f"{seed_name} (plant)"
                if plant_label not in reagent_plants[reagent_id]:
                    reagent_plants[reagent_id].append(plant_label)

    return dict(reagent_plants)


def build_all_sources(reagent_plants: dict, reaction_lookup: dict,
                      base_set: set) -> dict[str, list[str]]:
    """Build complete sources dict for each reagent."""
    sources = defaultdict(list)

    # 1. Plant sources
    for reagent_id, plants in reagent_plants.items():
        sources[reagent_id].extend(plants)

    # 2. Other known sources
    for reagent_id, src_list in OTHER_REAGENT_SOURCES.items():
        for src in src_list:
            if src not in sources[reagent_id]:
                sources[reagent_id].append(src)

    # 3. Dispenser chemicals
    for rid in base_set:
        if rid in BASE_DISPENSER_CHEMICALS:
            if "Chemical Dispenser" not in sources.get(rid, []):
                sources[rid].append("Chemical Dispenser")

    # 4. Chemistry reaction (mark crafted reagents)
    for rid in reaction_lookup:
        sources[rid].insert(0, "Chemistry reaction")

    return dict(sources)


# ─────────────────────────────────────────────
# Phase 9a: Antag Score Derivation
# ─────────────────────────────────────────────

# Auto-derive antag utility score from effectTags for reagents not in ANTAG_DATA.
# Capped at 5 so curated ANTAG_DATA entries (up to 10) always rank higher.
_ANTAG_TAG_WEIGHTS = {
    "deals:asphyxiation": 3, "deals:poison": 2, "deals:caustic": 2,
    "deals:heat": 1, "deals:cold": 1, "deals:radiation": 2,
    "deals:brute": 1, "deals:bloodloss": 2, "deals:shock": 1,
    "flammable": 2, "paralyze": 3, "explosion": 3, "stun": 2,
    "bleed": 2, "emote": 1, "speed": 1, "drunk": 0, "jitter": 0,
    "sleep": 3, "blind": 2, "mute": 2, "vomit": 1,
}

def compute_antag_score(effect_tags: list[str]) -> int:
    """Derive an antag utility score (0-5) from effect tags."""
    score = 0
    for tag in effect_tags:
        # Exact match first
        if tag in _ANTAG_TAG_WEIGHTS:
            score += _ANTAG_TAG_WEIGHTS[tag]
        # Prefix match (e.g., "deals:brute" matches "deals:")
        elif ":" in tag:
            prefix = tag.split(":")[0] + ":"
            for k, v in _ANTAG_TAG_WEIGHTS.items():
                if k.startswith(prefix):
                    score += max(1, v - 1)
                    break
    return min(score, 5)


# ─────────────────────────────────────────────
# Phase 9b: JSON Export
# ─────────────────────────────────────────────

def export_json(reagents: dict, reactions: dict, locale: dict,
                reaction_lookup: dict, base_set: set,
                all_sources: dict | None = None,
                fork_diffs: dict | None = None):
    """Export all data as a JSON file for the web frontend.
    fork_diffs: {fork_id: (blocked_set, modified_dict)} from auto-diff."""
    # Compute per-fork reagent counts
    fork_reagent_counts = {}
    for r in reagents.values():
        if r.get("abstract"):
            continue
        fid = detect_fork_source(r.get("_source_file", ""))
        fork_reagent_counts[fid] = fork_reagent_counts.get(fid, 0) + 1

    # Build fork metadata for frontend
    forks_meta = {}
    for fid, fconf in FORK_REGISTRY.items():
        count = fork_reagent_counts.get(fid, 0)
        if fid == "vanilla" or count > 0:
            forks_meta[fid] = {
                "name": fconf["name"],
                "color": fconf["color"],
                "reagentCount": count,
            }

    data = {
        "meta": {
            "generated": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "forks": forks_meta,
            # Backward compat
            "vanillaReagentCount": fork_reagent_counts.get("vanilla", 0),
            "rmcReagentCount": fork_reagent_counts.get("rmc14", 0),
            "reactionCount": len(reactions),
        },
        "reagents": {},
        "reactions": {},
        "baseChemicals": sorted(base_set),
        "categories": [],
        "edges": [],
    }

    # Build reagents
    cats_seen = set()
    for rid, reagent in sorted(reagents.items()):
        if reagent.get("abstract"):
            continue

        cat = categorize_reagent(reagent)
        cats_seen.add(cat)
        source = detect_fork_source(reagent.get("_source_file", ""))
        effects_str, metab_paths = summarize_metabolisms(reagent.get("metabolisms", {}))

        recipe_obj = None
        if rid in reaction_lookup:
            rxn = reaction_lookup[rid][0]
            reactants_obj = {}
            for react_id, info in rxn.get("reactants", {}).items():
                amount = info.get("amount", 1) if isinstance(info, dict) else 1
                catalyst = info.get("catalyst", False) if isinstance(info, dict) else False
                reactants_obj[react_id] = {"amount": amount, "catalyst": catalyst}
            recipe_obj = {
                "reactants": reactants_obj,
                "products": rxn.get("products", {}),
                "minTemp": rxn.get("minTemp"),
                "maxTemp": rxn.get("maxTemp"),
                "mixer": rxn.get("requiredMixerCategories", []),
            }

        # Overdose: try RMC14 top-level first, then vanilla threshold extraction
        overdose = None
        if reagent.get("overdose"):
            overdose = reagent["overdose"]
        else:
            metab_data = reagent.get("metabolisms", {})
            od = extract_overdose_threshold(metab_data)
            if od:
                overdose = od
        crit_overdose = reagent.get("criticalOverdose")

        # Effect tags for filtering
        effect_tags = extract_effect_tags(reagent.get("metabolisms", {}))

        # Antag mode data — curated or auto-derived
        antag_info = ANTAG_DATA.get(rid, {})
        antag_score = antag_info.get("score", 0) if antag_info else compute_antag_score(effect_tags)
        antag_tags = antag_info.get("tags", [])
        antag_tips = antag_info.get("tips", "")

        reagent_obj = {
            "id": rid,
            "name": resolve_name(reagent, locale),
            "group": reagent.get("group", ""),
            "category": cat,
            "source": source,
            "color": reagent.get("color", ""),
            "desc": resolve_desc(reagent, locale),
            "physicalDesc": locale.get(reagent.get("physicalDesc", ""), reagent.get("physicalDesc", "")),
            "flavor": str(reagent.get("flavor", "")),
            "isBase": rid in base_set,
            "isDispenser": rid in BASE_DISPENSER_CHEMICALS,
            "effects": effects_str,
            "effectTags": effect_tags,
            "metabolismPaths": metab_paths.split(", ") if metab_paths else [],
            "overdose": overdose,
            "criticalOverdose": crit_overdose,
            "recipe": recipe_obj,
            "obtainSources": (all_sources or {}).get(rid, []),
        }
        # Only include antag fields when non-default (saves ~30KB in JSON)
        if antag_score:
            reagent_obj["antagScore"] = antag_score
        if antag_tags:
            reagent_obj["antagTags"] = antag_tags
        if antag_tips:
            reagent_obj["antagTips"] = antag_tips

        data["reagents"][rid] = reagent_obj

    # Build reactions
    for rid, rxn in sorted(reactions.items()):
        source = detect_fork_source(rxn.get("_source_file", ""))
        reactants_obj = {}
        for react_id, info in rxn.get("reactants", {}).items():
            amount = info.get("amount", 1) if isinstance(info, dict) else 1
            catalyst = info.get("catalyst", False) if isinstance(info, dict) else False
            reactants_obj[react_id] = {"amount": amount, "catalyst": catalyst}

        # Per-fork availability status for vanilla reactions (uses auto-diff results)
        fork_status = {}
        fork_notes = {}
        if source == "vanilla" and fork_diffs:
            for fid in FORK_REGISTRY:
                if fid == "vanilla":
                    continue
                blocked, modified = fork_diffs.get(fid, (set(), {}))
                if rid in blocked:
                    fork_status[fid] = "blocked"
                elif rid in modified:
                    fork_status[fid] = "modified"
                    fork_notes[fid] = modified[rid]
                # "available" is default — omit to save space
        elif source == "vanilla":
            # Fallback: use manual config if fork_diffs not provided
            for fid, fconf in FORK_REGISTRY.items():
                if fid == "vanilla":
                    continue
                blocked = fconf.get("blocked_reactions", set())
                modified = fconf.get("modified_reactions", {})
                if rid in blocked:
                    fork_status[fid] = "blocked"
                elif rid in modified:
                    fork_status[fid] = "modified"
                    fork_notes[fid] = modified[rid]

        reaction_obj = {
            "id": rid,
            "source": source,
            "reactants": reactants_obj,
            "products": rxn.get("products", {}),
            "minTemp": rxn.get("minTemp"),
            "maxTemp": rxn.get("maxTemp"),
            "mixer": rxn.get("requiredMixerCategories", []),
            "effects": summarize_reaction_effects(rxn),
            "impact": rxn.get("impact"),
        }
        # Only add fork fields if there's something to report
        if fork_status:
            reaction_obj["forkStatus"] = fork_status
        if fork_notes:
            reaction_obj["forkNotes"] = fork_notes
        # Backward compat: rmcStatus / rmcNote
        rmc_status = fork_status.get("rmc14", "available")
        reaction_obj["rmcStatus"] = rmc_status
        if rmc_status == "modified":
            reaction_obj["rmcNote"] = fork_notes.get("rmc14", "")
        else:
            reaction_obj["rmcNote"] = ""

        data["reactions"][rid] = reaction_obj

    # Build edges for vis.js graph
    for rid, rxn in reactions.items():
        products = rxn.get("products", {})
        reactants = rxn.get("reactants", {})
        for product_id in products:
            for react_id, info in reactants.items():
                amount = info.get("amount", 1) if isinstance(info, dict) else 1
                catalyst = info.get("catalyst", False) if isinstance(info, dict) else False
                data["edges"].append({
                    "from": react_id,
                    "to": product_id,
                    "reaction": rid,
                    "amount": amount,
                    "catalyst": catalyst,
                })

    data["categories"] = sorted(cats_seen)
    data["warnings"] = DANGEROUS_INTERACTIONS

    # Antag mode data
    data["antagStrategies"] = ANTAG_STRATEGIES
    data["deliveryMechanisms"] = DELIVERY_MECHANISMS
    data["syndicateItems"] = SYNDICATE_ITEMS

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    print(f"  JSON saved to: {JSON_FILE}")
    print(f"  Reagents: {len(data['reagents'])}, Reactions: {len(data['reactions'])}, Edges: {len(data['edges'])}")
    return data


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main():
    loader = make_yaml_loader()

    # Phase 1: Fetch all forks
    print("=== Phase 1: Fetching files from GitHub ===")

    fork_data = {}  # fork_id -> {"reagent_files": {}, "reaction_files": {}, "locale_files": {}, "seed_files": {}}
    for fork_id, fconf in FORK_REGISTRY.items():
        url = fconf["raw_url"]

        print(f"\n[{fconf['name']}] Reagents...")
        reagent_files = fetch_all_files(fconf["reagent_files"], url, fork_id)
        print(f"  Fetched {len(reagent_files)} reagent files")

        print(f"[{fconf['name']}] Reactions...")
        reaction_files = fetch_all_files(fconf["reaction_files"], url, fork_id)
        print(f"  Fetched {len(reaction_files)} reaction files")

        print(f"[{fconf['name']}] Locale...")
        locale_files = fetch_all_files(fconf.get("locale_files", []), url, fork_id)
        print(f"  Fetched {len(locale_files)} locale files")

        seed_files = fetch_all_files(fconf.get("seed_files", []), url, fork_id)
        if seed_files:
            print(f"  Fetched {len(seed_files)} seed files")

        fork_data[fork_id] = {
            "reagent_files": reagent_files,
            "reaction_files": reaction_files,
            "locale_files": locale_files,
            "seed_files": seed_files,
        }

    # Phase 1b: Fetch vanilla-path files from each fork for auto-diff
    fork_vanilla_overrides = {}  # fork_id -> {path: content}
    for fork_id, fconf in FORK_REGISTRY.items():
        if fork_id == "vanilla":
            continue
        override_files = fconf.get("vanilla_override_reaction_files", [])
        if override_files:
            print(f"\n[{fconf['name']}] Vanilla-path overrides...")
            overrides = fetch_all_files(override_files, fconf["raw_url"], f"{fork_id}_vanilla_overrides")
            fork_vanilla_overrides[fork_id] = overrides
            print(f"  Fetched {len(overrides)} vanilla-path override files")

    # Botany locale (vanilla only for now)
    botany_locale_files = fetch_all_files(
        FORK_REGISTRY["vanilla"].get("botany_locale_files", []),
        FORK_REGISTRY["vanilla"]["raw_url"], "vanilla")
    print(f"  Botany locale: {len(botany_locale_files)} files")

    # Phase 2: Parse YAML
    print("\n=== Phase 2: Parsing YAML prototypes ===")

    parsed = {}  # fork_id -> {"reagents": {}, "reactions": {}}
    for fork_id, fdata in fork_data.items():
        reagents, reactions = parse_all_prototypes(fdata["reagent_files"], loader)
        r2, rxn2 = parse_all_prototypes(fdata["reaction_files"], loader)
        reactions.update(rxn2)
        parsed[fork_id] = {"reagents": reagents, "reactions": reactions}
        print(f"  {FORK_REGISTRY[fork_id]['name']}: {len(reagents)} reagents, {len(reactions)} reactions")

    # Phase 3: Localization
    print("\n=== Phase 3: Loading localization ===")
    locale = {}
    for fork_id, fdata in fork_data.items():
        locale.update(load_all_localization(fdata["locale_files"]))
    print(f"  Locale entries: {len(locale)}")

    # Phase 4: Merge & resolve parents
    print("\n=== Phase 4: Resolving parent inheritance ===")
    # Merge: vanilla first, then each fork layers on top (fork reagents override vanilla by ID)
    all_reagents = dict(parsed["vanilla"]["reagents"])
    all_reactions = dict(parsed["vanilla"]["reactions"])
    for fork_id, pdata in parsed.items():
        if fork_id == "vanilla":
            continue
        all_reagents.update(pdata["reagents"])
        all_reactions.update(pdata["reactions"])

    # Seed files — combine from all forks
    v_seed_files = fork_data.get("vanilla", {}).get("seed_files", {})

    all_reagents = resolve_parents(all_reagents)
    print(f"  Total reagents after resolution: {len(all_reagents)}")
    print(f"  Total reactions: {len(all_reactions)}")

    # Phase 4b: Auto-diff fork vanilla overrides against vanilla originals
    # (runs after Phase 4 so all_reagents is available for category-level blocking)
    print("\n=== Phase 4b: Auto-detecting fork reaction diffs ===")
    # Build a quick category lookup from resolved reagents
    reagent_categories = {}
    for rid, rdata in all_reagents.items():
        cat = categorize_reagent(rdata)
        reagent_categories[rid] = {"category": cat}

    fork_diffs = {}  # fork_id -> (blocked_set, modified_dict)
    for fork_id, override_files in fork_vanilla_overrides.items():
        if not override_files:
            continue
        _, fork_vanilla_rxns = parse_all_prototypes(override_files, loader)
        fconf = FORK_REGISTRY[fork_id]
        manual_blocked = fconf.get("blocked_reactions", set())
        manual_modified = fconf.get("modified_reactions", {})
        blocked_cats = fconf.get("blocked_categories", set())
        blocked, modified = diff_fork_reactions(
            parsed["vanilla"]["reactions"], fork_vanilla_rxns,
            manual_blocked, manual_modified,
            blocked_categories=blocked_cats,
            all_reagents=reagent_categories,
        )
        fork_diffs[fork_id] = (blocked, modified)
        auto_b = len(blocked) - len(manual_blocked)
        auto_m = len(modified) - len(manual_modified)
        print(f"  {fconf['name']}: {len(blocked)} blocked ({auto_b} auto), {len(modified)} modified ({auto_m} auto)")

    # For forks without override files, fall back to manual config
    for fork_id, fconf in FORK_REGISTRY.items():
        if fork_id == "vanilla" or fork_id in fork_diffs:
            continue
        fork_diffs[fork_id] = (
            fconf.get("blocked_reactions", set()),
            fconf.get("modified_reactions", {}),
        )

    # Phase 5: Build dependency trees
    print("\n=== Phase 5: Building craft dependency trees ===")
    reaction_lookup = build_reaction_lookup(all_reactions)
    base_set = identify_base_chemicals(all_reagents, reaction_lookup)
    print(f"  Products with recipes: {len(reaction_lookup)}")
    print(f"  Base chemicals: {len(base_set)}")

    # Phase 5b: Parse plant/seed sources
    print("\n=== Phase 5b: Parsing plant/seed sources ===")
    botany_locale = load_all_localization(botany_locale_files)
    all_seed_locale = {**locale, **botany_locale}
    all_seed_files = {}
    for fdata in fork_data.values():
        all_seed_files.update(fdata.get("seed_files", {}))
    reagent_plants = parse_seed_sources(all_seed_files, loader, all_seed_locale)
    print(f"  Plant reagent sources: {len(reagent_plants)}")
    all_sources = build_all_sources(reagent_plants, reaction_lookup, base_set)
    print(f"  Total reagents with sources: {len(all_sources)}")

    # Phase 6-7: Generate Excel
    print("\n=== Phase 6-7: Generating Excel ===")
    wb = generate_excel(all_reagents, all_reactions, locale, reaction_lookup, base_set)
    wb.save(str(OUTPUT_FILE))
    print(f"\n  Saved to: {OUTPUT_FILE}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"    {name}: {ws.max_row - 1 if ws.max_row > 1 else 0} rows")

    # Phase 8: Generate JSON for web frontend
    print("\n=== Phase 8: Generating JSON for web frontend ===")
    export_json(all_reagents, all_reactions, locale, reaction_lookup, base_set, all_sources, fork_diffs)

    print("\n=== Phase 9: Extracting sprites from SS14 repo ===")
    fetch_and_extract_sprites()

    print("\nDone!")


if __name__ == "__main__":
    main()
